#!/usr/bin/env python3
"""
Swiss Timing / World Boxing PDF -> structured tables.

Extracts a competition results PDF (the "Session Results", "Medallists by Weight
Category", "Top List", "Medal Standings" and "Ranking by Team" reports produced by
Swiss Timing) into CSVs and/or a formatted Excel workbook.

No OCR. These PDFs are born-digital, so the text layer is read directly and every
character is exact. Parsing is POSITIONAL: each report has fixed column x-ranges,
which survives long names, blank cells and wrapped rows far better than regex on
flattened text.

Usage:
    python boxam_extract.py results.pdf -o out/              # CSVs + xlsx
    python boxam_extract.py results.pdf -o out/ --no-excel   # CSVs only
    python boxam_extract.py results.pdf --check              # parse + report only

Requires: pdfplumber   (openpyxl only if writing Excel)
"""

import argparse
import collections
import csv
import os
import re
import sys

import pdfplumber

# --------------------------------------------------------------------------
# Column geometry of the "Session Results" report, in PDF points.
# Read off the header row; stable across every session page.
# --------------------------------------------------------------------------
COLS = {
    'order':    (40, 60),
    'bout':     (60, 82),
    'weight':   (82, 165),
    'refflag':  (165, 196),
    'corner':   (196, 235),
    'name':     (235, 342),
    'judgelbl': (342, 373),
    'team':     (373, 415),
    'winner':   (415, 475),
    'result':   (475, 510),
    'decision': (510, 533),
    'score':    (533, 600),
}

# Judge names overflow their nominal column, so read them to the score edge.
JUDGE_NAME_SPAN = (415, 533)

LINE_TOL = 3        # points; words within this vertical distance are one line
NUMERIC = re.compile(r'^-?\d+(\.\d+)?$')


# --------------------------------------------------------------------------
# helpers
# --------------------------------------------------------------------------
def span(words, x0, x1):
    """Join the words whose left edge falls inside [x0, x1)."""
    return ' '.join(w['text'] for w in words if x0 <= w['x0'] < x1).strip()


def cell(words, key):
    return span(words, *COLS[key])


def group_lines(page, tol=LINE_TOL):
    """Cluster a page's words into visual lines, each sorted left to right."""
    buckets = []
    for w in sorted(page.extract_words(), key=lambda z: (z['top'], z['x0'])):
        for b in buckets:
            if abs(b[0]['top'] - w['top']) <= tol:
                b.append(w)
                break
        else:
            buckets.append([w])
    return [sorted(b, key=lambda z: z['x0']) for b in buckets]


def coerce(v):
    """'12' -> 12, '2.4' -> 2.4, True -> 'YES', everything else unchanged."""
    if isinstance(v, str) and NUMERIC.match(v.strip()):
        s = v.strip()
        return float(s) if '.' in s else int(s)
    if v is True:
        return 'YES'
    if v is False:
        return ''
    return v


def ruled_rows(page):
    """Flatten every ruled table on a page into a list of cleaned string rows."""
    out = []
    for table in page.extract_tables():
        for row in table:
            out.append([(c or '').replace('\n', ' ').strip() for c in row])
    return out


# --------------------------------------------------------------------------
# 1. Session Results -> bouts + judge scorecards
# --------------------------------------------------------------------------
def parse_sessions(pdf):
    bouts, scorecards = [], []
    current_ring = None

    for pageno, page in enumerate(pdf.pages, 1):
        text = page.extract_text() or ''
        if 'Session Results' not in text:
            continue

        m = re.search(r'Session\s+(\d+)', text)
        session = int(m.group(1)) if m else None
        m = re.search(r'([A-ZÁ]{2,3}\.\s+\d+\s+\w+\.\s+\d{4})\s+(\d{2}:\d{2})', text)
        date, start = (m.group(1), m.group(2)) if m else (None, None)
        m = re.search(r'RING\s*-?\s*([AB])', text)
        if m:                              # ring is printed once per report,
            current_ring = 'RING ' + m.group(1)   # so carry it forward

        lines = group_lines(page)
        i = 0
        while i < len(lines):
            head = lines[i]
            order, weight = cell(head, 'order'), cell(head, 'weight')

            # A bout starts on the line carrying both an order number and a
            # weight category. Everything else on the page is chrome.
            if not (order.isdigit() and weight.startswith('Elite')):
                i += 1
                continue

            red = {
                'name': cell(head, 'name'), 'team': cell(head, 'team'),
                'winner': cell(head, 'winner'), 'result': cell(head, 'result'),
                'decision': cell(head, 'decision'),
            }
            bout_no = int(cell(head, 'bout'))

            blue = {'name': '', 'team': '', 'winner': '', 'result': '', 'decision': ''}
            j = i + 1
            if j < len(lines) and cell(lines[j], 'corner') == 'BLUE':
                b = lines[j]
                blue = {
                    'name': cell(b, 'name'), 'team': cell(b, 'team'),
                    'winner': cell(b, 'winner'), 'result': cell(b, 'result'),
                    'decision': cell(b, 'decision'),
                }
                j += 1

            # The winner/result/decision block is printed on whichever corner won.
            winner_corner = red['winner'] or blue['winner']
            result = red['result'] or blue['result']
            decision = red['decision'] or blue['decision']

            # Officials block: one referee line, then up to five judge lines.
            ref_team = ref_name = ''
            judges = []
            while j < len(lines):
                ln = lines[j]
                flat = ' '.join(w['text'] for w in ln)

                if cell(ln, 'order').isdigit() and cell(ln, 'weight').startswith('Elite'):
                    break                                   # next bout
                if 'Referee:' in flat:
                    ref_team, ref_name = cell(ln, 'corner'), cell(ln, 'name')

                jm = re.match(r'Judge\s*(\d):', cell(ln, 'judgelbl'))
                if jm:
                    raw = span(ln, *JUDGE_NAME_SPAN)
                    # A trailing '*' is the "preferred winner" marker on a drawn
                    # card; it belongs to the score, not the judge's name.
                    star = raw.endswith('*')
                    judges.append({
                        'no': int(jm.group(1)),
                        'team': cell(ln, 'team'),
                        'name': re.sub(r'\s*\*$', '', raw),
                        'score': ('* ' if star else '') + cell(ln, 'score'),
                    })
                elif 'Referee:' not in flat and cell(ln, 'corner') in ('RED', 'BLUE'):
                    break
                j += 1

            wm = re.match(r'Elite (Men|Women)\s*([+-])?\s*(\d+)\s*Kg\s*\((\S+?)\)', weight)
            gender = wm.group(1) if wm else ''
            wcode = wm.group(4) if wm else ''

            side = red if winner_corner == 'RED' else (blue if winner_corner == 'BLUE' else None)

            rec = {
                'Session': session, 'Date': date, 'Start Time': start,
                'Ring': current_ring, 'Page': pageno,
                'Order': int(order), 'Bout No': bout_no,
                'Gender': gender, 'Weight Category': weight, 'Weight Code': wcode,
                'Red Boxer': red['name'], 'Red Team': red['team'],
                'Blue Boxer': blue['name'], 'Blue Team': blue['team'],
                'Winner Corner': winner_corner,
                'Winner Boxer': side['name'] if side else '',
                'Winner Team': side['team'] if side else '',
                'Result': result, 'Decision': decision,
                'Referee Team': ref_team, 'Referee Name': ref_name,
            }
            for k in range(1, 6):
                jd = next((x for x in judges if x['no'] == k), None)
                rec[f'Judge {k} Team'] = jd['team'] if jd else ''
                rec[f'Judge {k} Name'] = jd['name'] if jd else ''
                rec[f'Judge {k} Score'] = jd['score'] if jd else ''
            bouts.append(rec)

            for jd in judges:
                sm = re.match(r'^(\d+):(\d+)$', jd['score'].replace('*', '').strip())
                scorecards.append({
                    'Session': session, 'Bout No': bout_no, 'Weight Category': weight,
                    'Red Boxer': red['name'], 'Blue Boxer': blue['name'],
                    'Judge No': jd['no'], 'Judge Team': jd['team'], 'Judge Name': jd['name'],
                    'Score': jd['score'],
                    'Red Points': int(sm.group(1)) if sm else '',
                    'Blue Points': int(sm.group(2)) if sm else '',
                    'Preferred Winner Flag': '*' in jd['score'],
                })
            i = j

    return bouts, scorecards


# --------------------------------------------------------------------------
# 2. Medallists by Weight Category  (ruled table)
# --------------------------------------------------------------------------
def parse_medallists(pdf):
    rows, category = [], ''
    for page in pdf.pages:
        if 'Medallists by Weight Category' not in (page.extract_text() or ''):
            continue
        for r in ruled_rows(page):
            if not r or r[0].startswith('Weight Category'):
                continue
            if r[0]:
                category = r[0]                       # only on the first of 4 rows
            if len(r) >= 5 and r[2] in ('GOLD', 'SILVER', 'BRONZE'):
                rows.append({'Weight Category': category, 'Date': r[1], 'Medal': r[2],
                             'Boxer': r[3], 'Team Code': r[4]})
    return rows


# --------------------------------------------------------------------------
# 3. Top List / final rankings  (two side-by-side blocks, unruled -> positional)
# --------------------------------------------------------------------------
def parse_toplist(pdf):
    rows = []
    blocks = {'L': (25, 300), 'R': (300, 600)}
    for page in pdf.pages:
        if 'Top List' not in (page.extract_text() or ''):
            continue
        by_top = collections.defaultdict(list)
        for w in page.extract_words():
            by_top[round(w['top'])].append(w)

        category = {'L': '', 'R': ''}
        for top in sorted(by_top):
            line = sorted(by_top[top], key=lambda z: z['x0'])
            for side, (lo, hi) in blocks.items():
                col = [w for w in line if lo <= w['x0'] < hi]
                if not col:
                    continue
                text = ' '.join(w['text'] for w in col)
                if text.startswith('Elite'):
                    category[side] = text
                    continue
                rank = [w for w in col if w['text'].isdigit() and w['x0'] < lo + 25]
                if not rank:
                    continue
                name = span(col, lo + 25, lo + 195)
                seed = span(col, lo + 195, lo + 220)
                team = span(col, lo + 220, hi)
                if name and team:
                    rows.append({'Weight Category': category[side],
                                 'Rank': int(rank[0]['text']), 'Boxer': name,
                                 'Seed': seed, 'Team Code': team})
    return rows


# --------------------------------------------------------------------------
# 4 & 5. Medal Standings and Ranking by Team  (ruled tables)
# --------------------------------------------------------------------------
MEDAL_COLS = ['Rank', 'Team', 'Men Gold', 'Men Silver', 'Men Bronze', 'Men Total',
              'Women Gold', 'Women Silver', 'Women Bronze', 'Women Total',
              'Total Gold', 'Total Silver', 'Total Bronze', 'Total Medals',
              'Rank by Total']

TEAM_COLS = ['Rank', 'Team Code', 'Boxers', 'Wins Prelims', 'Wins Semifinals',
             'Wins Finals', 'Lost', 'Bouts', 'Bouts per Boxer', 'Points',
             'Points per Boxer', 'Gold', 'Silver', 'Bronze', 'Total Medals']


def parse_grid(pdf, marker, columns):
    rows = []
    for page in pdf.pages:
        if marker not in (page.extract_text() or ''):
            continue
        for r in ruled_rows(page):
            if len(r) < len(columns) - 1 or not r[0].strip().isdigit():
                continue
            vals = [c if c else '0' for c in r]
            vals += ['0'] * (len(columns) - len(vals))
            rec = dict(zip(columns, vals[:len(columns)]))
            rec['Rank'] = int(rec['Rank'])
            rows.append(rec)
    return rows


# --------------------------------------------------------------------------
# writers
# --------------------------------------------------------------------------
def write_csv(rows, path):
    if not rows:
        return
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        for r in rows:
            w.writerow({k: coerce(v) for k, v in r.items()})


def write_excel(tables, path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    hdr_fill = PatternFill('solid', fgColor='1F3864')
    hdr_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    body = Font(name='Arial', size=10)

    wb = Workbook()
    wb.remove(wb.active)
    for title, rows in tables.items():
        if not rows:
            continue
        ws = wb.create_sheet(title[:31])
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([coerce(r.get(h, '')) for h in headers])
        for c in ws[1]:
            c.fill, c.font = hdr_fill, hdr_font
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.font = body
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        ws.row_dimensions[1].height = 30
        for i, h in enumerate(headers, 1):
            widest = max([len(str(h))] + [len(str(coerce(r.get(h, '')))) for r in rows]) + 2
            ws.column_dimensions[get_column_letter(i)].width = min(max(widest, 9), 34)
    wb.save(path)


# --------------------------------------------------------------------------
def integrity_report(bouts, scorecards, medallists):
    """Sanity checks worth eyeballing before trusting the output."""
    print(f'  bouts parsed          : {len(bouts)}')
    print(f'  judge scorecard rows  : {len(scorecards)}')
    print(f'  medallist rows        : {len(medallists)}')

    nums = sorted(b['Bout No'] for b in bouts)
    if nums:
        gaps = sorted(set(range(nums[0], nums[-1] + 1)) - set(nums))
        print(f'  bout numbers          : {nums[0]}-{nums[-1]}'
              + (f'  MISSING: {gaps}' if gaps else '  (contiguous)'))
    print('  results               : '
          + ', '.join(f'{k}={v}' for k, v in
                      sorted(collections.Counter(b['Result'] for b in bouts).items())))

    for label, key in (('bouts with no winner', 'Winner Corner'),
                       ('bouts with no blue boxer', 'Blue Boxer'),
                       ('bouts with no red boxer', 'Red Boxer')):
        bad = [b['Bout No'] for b in bouts if not b[key]]
        if bad:
            print(f'  WARNING {label}: {bad}')

    names = {b[f'Judge {k} Name'] for b in bouts for k in range(1, 6)} - {''}
    print(f'  distinct judges       : {len(names)}')
    print('  NOTE: near-identical official names may be source typos '
          '(e.g. PARKHATOV / PARAKHATOV). Normalise before grouping.')


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('pdf', help='path to the Swiss Timing results PDF')
    ap.add_argument('-o', '--outdir', default='.', help='output directory')
    ap.add_argument('--no-excel', action='store_true', help='write CSVs only')
    ap.add_argument('--check', action='store_true',
                    help='parse and print the integrity report, write nothing')
    args = ap.parse_args()

    if not os.path.exists(args.pdf):
        sys.exit(f'error: no such file: {args.pdf}')

    with pdfplumber.open(args.pdf) as pdf:
        bouts, scorecards = parse_sessions(pdf)
        medallists = parse_medallists(pdf)
        toplist = parse_toplist(pdf)
        standings = parse_grid(pdf, 'Medal Standings', MEDAL_COLS)
        teamrank = parse_grid(pdf, 'Ranking by Team', TEAM_COLS)

    tables = {
        'Bouts': bouts,
        'Judge_Scorecards': scorecards,
        'Medallists': medallists,
        'Final_Rankings': toplist,
        'Medal_Standings': standings,
        'Team_Ranking': teamrank,
    }

    print(f'Parsed {args.pdf}')
    integrity_report(bouts, scorecards, medallists)

    if args.check:
        return

    os.makedirs(args.outdir, exist_ok=True)
    for title, rows in tables.items():
        path = os.path.join(args.outdir, f'{title.lower()}.csv')
        write_csv(rows, path)
        if rows:
            print(f'  wrote {path}  ({len(rows)} rows)')

    if not args.no_excel:
        xlsx = os.path.join(args.outdir, 'results_structured.xlsx')
        write_excel(tables, xlsx)
        print(f'  wrote {xlsx}')


if __name__ == '__main__':
    main()